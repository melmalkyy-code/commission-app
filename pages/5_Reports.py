import sys, os, io, base64; sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
import streamlit as st
import pandas as pd
from src.startup import init_db
from src.auth import require_login
init_db()
require_login()
from src.models import get_setting, get_or_create_period, get_period, get_branches, get_salespersons
from src.calculations import calc_all_commissions, get_totals

from src.ui import inject_css, page_header, sidebar_logo, render_df
from src.i18n import t, tl, q_label, is_rtl
from src.pdf_arabic import cell as _ar_cell, pdf_font, pdf_font_bold, ar as _ar
from src.branding import page_icon

PRIMARY = get_setting('primary_color', '#354f61')
ACCENT  = get_setting('accent_color', '#f6ba3b')
COMPANY = get_setting('company_name', 'Surveying Experts')
_LOGO_B64   = get_setting('company_logo_b64', '')
_LOGO_BYTES = base64.b64decode(_LOGO_B64) if _LOGO_B64 else None
st.set_page_config(page_title="Reports Center — Surveying Experts", page_icon=page_icon(), layout="wide")
inject_css(PRIMARY)
sidebar_logo(COMPANY, PRIMARY)
page_header(t("Reports Center"), t("Download commission reports at company, branch, and salesperson level"), PRIMARY)

col1, col2, _ = st.columns([1, 1, 4])
year    = col1.selectbox(t("Year"),    [2024, 2025, 2026, 2027], index=2, key="rep_year")
quarter = col2.selectbox(t("Quarter"), [1, 2, 3, 4],             index=1, key="rep_q",
                          format_func=q_label)
period  = get_or_create_period(year, quarter)
period_label = f"Q{quarter} {year}"

# ── Previous quarter for QoQ ─────────────────────────────────────────────────
_pq_y = year - 1 if quarter == 1 else year
_pq_q = 4        if quarter == 1 else quarter - 1
_prev_period = get_period(_pq_y, _pq_q)   # SELECT only — never inserts
_prev_label  = f"Q{_pq_q} {_pq_y}"

st.divider()

_EMPTY_T = {'total_sales': 0, 'total_target': 0, 'achievement': 0,
            'total_base': 0, 'total_final': 0, 'achieved_count': 0, 'total_count': 0}

_all_branches    = get_branches()
_branch_to_region = {b['name']: (b.get('region') or '') for b in _all_branches}

with st.spinner(t("Loading data...")):
    commissions = calc_all_commissions(period['id'])
    totals      = get_totals(commissions)
    if _prev_period:
        _prev_commissions = calc_all_commissions(_prev_period['id'])
        _prev_totals      = get_totals(_prev_commissions)
    else:
        _prev_commissions = []
        _prev_totals      = _EMPTY_T
    _prev_sp_map = {c['salesperson_name']: c for c in _prev_commissions}


def _qoq_g(curr: float, prev: float) -> str:
    if prev == 0:
        return "—"
    d = (curr - prev) / prev * 100
    return f"{'▲' if d >= 0 else '▼'} {abs(d):.1f}%"


# ─── RTL helpers for Arabic PDF output ───────────────────────────────────────
def _is_ar(lang) -> bool:
    return lang == 'ar'


def _rtl_rows(data, lang):
    """Reverse each row's columns so the first logical column sits on the right."""
    return [list(reversed(r)) for r in data] if _is_ar(lang) else data


def _rtl_widths(widths, lang):
    return list(reversed(widths)) if (widths and _is_ar(lang)) else widths


def _rtl_halign(lang) -> str:
    return 'RIGHT' if _is_ar(lang) else 'LEFT'


def _rtl_style(lang):
    """Extra TableStyle commands to right-align every cell for Arabic."""
    return [('ALIGN', (0, 0), (-1, -1), 'RIGHT')] if _is_ar(lang) else []


def _para_align(lang):
    from reportlab.lib.enums import TA_RIGHT, TA_LEFT
    return TA_RIGHT if _is_ar(lang) else TA_LEFT


# ─── Excel helpers ────────────────────────────────────────────────────────────
def _fill_row(ws, row_idx, fill_hex, font_hex="FFFFFF", bold=True):
    from openpyxl.styles import PatternFill, Font, Alignment
    fill = PatternFill(start_color=fill_hex.lstrip('#'), end_color=fill_hex.lstrip('#'), fill_type="solid")
    font = Font(color=font_hex, bold=bold)
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


def _insert_logo_xl(ws):
    if not _LOGO_BYTES:
        return
    try:
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(io.BytesIO(_LOGO_BYTES))
        img.width, img.height = 90, 45
        ws.add_image(img, 'A1')
        ws.row_dimensions[1].height = 40
    except Exception:
        pass


# ─── Company Excel ────────────────────────────────────────────────────────────
def make_excel_company(lang: str = 'en') -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ph = PRIMARY.lstrip('#')
    def _t(s): return tl(s, lang)
    _rtl = lang == 'ar'

    # Sheet 1: Summary
    ws = wb.active
    ws.title = _t("Company Summary")
    if _rtl: ws.sheet_view.rightToLeft = True
    ws.append([COMPANY, "", f"{_t('Commission Report')} - {period_label}"])
    _fill_row(ws, 1, PRIMARY)
    _insert_logo_xl(ws)
    ws.append([])
    ws.append([_t("Metric"), _t("Value")])
    _fill_row(ws, 3, PRIMARY)
    for row in [
        (_t("Total Sales (SAR)"),      round(totals['total_sales'], 0)),
        (_t("Total Target (SAR)"),     round(totals['total_target'], 0)),
        (_t("Achievement %"),          round(totals['achievement'], 1)),
        (_t("Base Commission (SAR)"),  round(totals['total_base'], 0)),
        (_t("Final Commission (SAR)"), round(totals['total_final'], 0)),
        (_t("Total Salespersons"),     totals['total_count']),
        (_t("Achieved Target"),        totals['achieved_count']),
    ]:
        ws.append(list(row))
    _auto_width(ws)

    # Sheet 2: All Salespersons
    ws2 = wb.create_sheet(_t("All Salespersons"))
    if _rtl: ws2.sheet_view.rightToLeft = True
    ws2.append([_t("Salesperson"), _t("Branch"), _t("Tier"), _t("Total Sales"), _t("Target"),
                _t("Achievement %"), _t("Base Comm."), _t("KPI Score"),
                _t("KPI Multiplier"), _t("Final Comm.")])
    _fill_row(ws2, 1, PRIMARY)
    for c in commissions:
        ws2.append([c['salesperson_name'], c['branch_name'], c['tier_name'],
                    round(c['total_actual'], 0), round(c['total_target'], 0),
                    round(c['achievement'], 1), round(c['base_commission'], 0),
                    round(c['kpi_score'], 2), c['kpi_multiplier'],
                    round(c['final_commission'], 0)])
    _auto_width(ws2)

    # Sheet 3: By Region
    ws3r = wb.create_sheet(_t("By Region"))
    if _rtl: ws3r.sheet_view.rightToLeft = True
    ws3r.append([_t("Region"), _t("Salespersons"), _t("Total Sales"), _t("Target"),
                 _t("Achievement %"), _t("Base Comm."), _t("Final Comm.")])
    _fill_row(ws3r, 1, PRIMARY)
    region_map_xl: dict = {}
    for c in commissions:
        reg = _branch_to_region.get(c['branch_name'] or '', '') or _t("No Region")
        if reg not in region_map_xl:
            region_map_xl[reg] = {'count': 0, 'sales': 0, 'target': 0, 'base': 0, 'final': 0}
        region_map_xl[reg]['count'] += 1
        region_map_xl[reg]['sales']  += c['total_actual']
        region_map_xl[reg]['target'] += c['total_target']
        region_map_xl[reg]['base']   += c['base_commission']
        region_map_xl[reg]['final']  += c['final_commission']
    for reg, rv in region_map_xl.items():
        ach = (rv['sales'] / rv['target'] * 100) if rv['target'] else 0
        ws3r.append([reg, rv['count'], round(rv['sales'], 0), round(rv['target'], 0),
                     round(ach, 1), round(rv['base'], 0), round(rv['final'], 0)])
    _auto_width(ws3r)

    # Sheet 4: By Branch
    ws3 = wb.create_sheet(_t("By Branch"))
    if _rtl: ws3.sheet_view.rightToLeft = True
    ws3.append([_t("Branch"), _t("Region"), _t("Salespersons"), _t("Total Sales"), _t("Target"),
                _t("Achievement %"), _t("Base Comm."), _t("Final Comm.")])
    _fill_row(ws3, 1, PRIMARY)
    branch_map: dict = {}
    for c in commissions:
        b = c['branch_name'] or 'Unknown'
        if b not in branch_map:
            branch_map[b] = {'count': 0, 'sales': 0, 'target': 0, 'base': 0, 'final': 0,
                             'region': _branch_to_region.get(b, '')}
        branch_map[b]['count'] += 1
        branch_map[b]['sales']  += c['total_actual']
        branch_map[b]['target'] += c['total_target']
        branch_map[b]['base']   += c['base_commission']
        branch_map[b]['final']  += c['final_commission']
    for br, bv in branch_map.items():
        ach = (bv['sales'] / bv['target'] * 100) if bv['target'] else 0
        ws3.append([br, bv.get('region', ''), bv['count'], round(bv['sales'], 0),
                    round(bv['target'], 0), round(ach, 1), round(bv['base'], 0),
                    round(bv['final'], 0)])
    _auto_width(ws3)

    # Sheet 4: By Category
    ws4 = wb.create_sheet(_t("By Category"))
    if _rtl: ws4.sheet_view.rightToLeft = True
    ws4.append([_t("Branch"), _t("Salesperson"), _t("Category"), _t("Actual Sales"),
                _t("Target"), _t("Achievement %"), _t("Rate %"), _t("Commission")])
    _fill_row(ws4, 1, PRIMARY)
    for c in commissions:
        for cr in c['categories']:
            ach = (cr['actual_sales'] / cr['target'] * 100) if cr['target'] else 0
            ws4.append([c['branch_name'], c['salesperson_name'], cr['category_name'],
                        round(cr['actual_sales'], 0), round(cr['target'], 0),
                        round(ach, 1), round(cr['rate'], 2), round(cr['commission'], 0)])
    _auto_width(ws4)

    # Sheet 5: QoQ Analysis
    if _prev_commissions:
        ws5 = wb.create_sheet(f"{_t('QoQ vs')} {_prev_label}")
        if _rtl: ws5.sheet_view.rightToLeft = True
        ws5.append([
            _t("Salesperson"), _t("Branch"),
            f"{_t('Sales')} {_prev_label}", f"{_t('Sales')} {period_label}", _t("Sales Growth"),
            f"{_t('Final Comm.')} {_prev_label}", f"{_t('Final Comm.')} {period_label}", _t("Comm. Growth"),
            f"{_t('Ach.')} {_prev_label}", f"{_t('Ach.')} {period_label}", _t("Ach. Δ pp"),
        ])
        _fill_row(ws5, 1, PRIMARY)
        for c in commissions:
            p  = _prev_sp_map.get(c['salesperson_name'], {})
            ps = p.get('total_actual', 0)
            pc = p.get('final_commission', 0)
            pa = p.get('achievement', 0)
            ws5.append([
                c['salesperson_name'], c['branch_name'],
                round(ps, 0), round(c['total_actual'], 0), _qoq_g(c['total_actual'], ps),
                round(pc, 0), round(c['final_commission'], 0), _qoq_g(c['final_commission'], pc),
                round(pa, 1), round(c['achievement'], 1),
                round(c['achievement'] - pa, 1),
            ])
        ws5.append([])
        ws5.append([
            _t("COMPANY TOTAL"), "",
            round(_prev_totals['total_sales'], 0), round(totals['total_sales'], 0),
            _qoq_g(totals['total_sales'], _prev_totals['total_sales']),
            round(_prev_totals['total_final'], 0), round(totals['total_final'], 0),
            _qoq_g(totals['total_final'], _prev_totals['total_final']),
            round(_prev_totals['achievement'], 1), round(totals['achievement'], 1),
            round(totals['achievement'] - _prev_totals['achievement'], 1),
        ])
        _fill_row(ws5, ws5.max_row, ACCENT, font_hex=ph, bold=True)
        _auto_width(ws5)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Branch Excel ─────────────────────────────────────────────────────────────
def make_excel_branch(branch_name: str, persons: list, lang: str = 'en') -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    def _t(s): return tl(s, lang)
    ws.title = _t("Branch Report")
    if lang == 'ar': ws.sheet_view.rightToLeft = True
    ws.append([COMPANY, branch_name, period_label])
    _fill_row(ws, 1, PRIMARY)
    _insert_logo_xl(ws)
    ws.append([])
    ws.append([_t("Salesperson"), _t("Tier"), _t("Sales"), _t("Target"), _t("Achievement %"),
               _t("Base Comm."), _t("KPI Score"), _t("KPI x"), _t("Final Comm."),
               f"{_t('Sales')} {_prev_label}", _t("Sales Growth"),
               f"{_t('Final Comm.')} {_prev_label}", _t("Comm. Growth")])
    _fill_row(ws, 3, PRIMARY)
    for c in persons:
        p  = _prev_sp_map.get(c['salesperson_name'], {})
        ps = p.get('total_actual', 0)
        pc = p.get('final_commission', 0)
        ws.append([
            c['salesperson_name'], c['tier_name'],
            round(c['total_actual'], 0), round(c['total_target'], 0),
            round(c['achievement'], 1), round(c['base_commission'], 0),
            round(c['kpi_score'], 2), c['kpi_multiplier'],
            round(c['final_commission'], 0),
            round(ps, 0), _qoq_g(c['total_actual'], ps),
            round(pc, 0), _qoq_g(c['final_commission'], pc),
        ])
    # Totals row
    ws.append([])
    b_sales  = sum(c['total_actual'] for c in persons)
    b_target = sum(c['total_target'] for c in persons)
    b_final  = sum(c['final_commission'] for c in persons)
    pb_sales = sum(_prev_sp_map.get(c['salesperson_name'], {}).get('total_actual', 0) for c in persons)
    pb_final = sum(_prev_sp_map.get(c['salesperson_name'], {}).get('final_commission', 0) for c in persons)
    ws.append([_t("TOTAL"), "", round(b_sales, 0), round(b_target, 0), "", "", "", "",
               round(b_final, 0),
               round(pb_sales, 0), _qoq_g(b_sales, pb_sales),
               round(pb_final, 0), _qoq_g(b_final, pb_final)])
    _fill_row(ws, ws.max_row, ACCENT, font_hex=PRIMARY.lstrip('#'), bold=True)
    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Salesperson Excel ────────────────────────────────────────────────────────
def make_excel_salesperson(c: dict, lang: str = 'en') -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    def _t(s): return tl(s, lang)
    ws.title = _t("Commission Report")
    if lang == 'ar': ws.sheet_view.rightToLeft = True
    ws.append([COMPANY, period_label, c['salesperson_name'], c['branch_name'], c['tier_name']])
    _fill_row(ws, 1, PRIMARY)
    _insert_logo_xl(ws)
    ws.append([])
    ws.append([_t("Category"), _t("Actual Sales"), _t("Target"), _t("Achievement %"),
               _t("Bracket"), _t("Rate %"), _t("Commission")])
    _fill_row(ws, 3, PRIMARY)
    for cr in c['categories']:
        ach = (cr['actual_sales'] / cr['target'] * 100) if cr['target'] else 0
        ws.append([cr['category_name'], round(cr['actual_sales'], 0),
                   round(cr['target'], 0), round(ach, 1),
                   cr['bracket'], round(cr['rate'], 2), round(cr['commission'], 0)])

    # KPI breakdown
    kpi = c.get('kpi') or {}
    kitems = kpi.get('items') or []
    ws.append([])
    ws.append([_t("KPI Score Breakdown")])
    _fill_row(ws, ws.max_row, PRIMARY)
    if not kitems:
        ws.append([_t("No KPI items configured.")])
    else:
        ws.append([_t("KPI Item"), _t("Weight %"), _t("Score"), _t("Max Score"),
                   _t("Achieved %"), _t("Weighted Points"), _t("Source")])
        _fill_row(ws, ws.max_row, PRIMARY)
        for it in kitems:
            ws.append([it['name'], it['weight'], round(it['raw_score'], 1),
                       it['max_score'], round(it['normalized'], 1),
                       round(it['contribution'], 1),
                       _t("Auto") if it.get('is_auto') else _t("Manual")])
        ws.append([_t("Weighted Score"), "", "", "", "",
                   round(kpi.get('weighted_score', 0), 1), ""])
        if kpi.get('bonus'):
            ws.append([_t("Bonus"), "", "", "", "", round(kpi['bonus'], 1), ""])
        if kpi.get('penalty'):
            ws.append([_t("Penalty"), "", "", "", "", -round(kpi['penalty'], 1), ""])
        ws.append([_t("Final KPI Score"), "", "", "", "",
                   round(kpi.get('final_score', c['kpi_score']), 1), ""])
        _rule = kpi.get('applied_rule')
        if _rule:
            _hi = "∞" if (_rule.get('is_unlimited') or _rule.get('score_to') is None) else f"{_rule['score_to']:.0f}"
            ws.append([_t("KPI Multiplier"), f"{_rule['score_from']:.0f}-{_hi}",
                       "", "", "", f"x {c['kpi_multiplier']}", ""])
        else:
            ws.append([_t("KPI Multiplier"), "", "", "", "", f"x {c['kpi_multiplier']}", ""])
        _fill_row(ws, ws.max_row, ACCENT, font_hex=PRIMARY.lstrip('#'), bold=True)

    ws.append([])
    for label, val in [(_t("Base Commission"), round(c['base_commission'], 0)),
                       (_t("KPI Score"),        round(c['kpi_score'], 2)),
                       (_t("KPI Multiplier"),   c['kpi_multiplier']),
                       (_t("FINAL COMMISSION"), round(c['final_commission'], 0))]:
        ws.append([label, val])
    _fill_row(ws, ws.max_row, ACCENT, font_hex=PRIMARY.lstrip('#'), bold=True)

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Letterhead helpers ──────────────────────────────────────────────────────
_LH_PATH = os.path.join(os.path.dirname(__file__), '..', 'assets', 'letterhead.png')


def _draw_page_bg(canvas, doc):
    """Draw the company letterhead as full-page background on every page."""
    from reportlab.lib.pagesizes import A4
    w, h = A4
    canvas.saveState()
    if os.path.exists(_LH_PATH):
        canvas.drawImage(_LH_PATH, 0, 0, width=w, height=h,
                         preserveAspectRatio=False)
    canvas.setFont('Helvetica', 7)
    canvas.setFillColorRGB(0.4, 0.4, 0.4)
    canvas.drawCentredString(w / 2, 22, f"Page {doc.page}")
    canvas.restoreState()


# ─── Company PDF ─────────────────────────────────────────────────────────────
def make_pdf_company(lang: str = 'en') -> bytes:
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.colors import HexColor
    from reportlab.lib.pagesizes import A4
    def _t(s): return tl(s, lang)
    _c = lambda s: _ar_cell(s, lang)
    _fn  = pdf_font(lang)
    _fnb = pdf_font_bold(lang)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=50, rightMargin=50,
                            topMargin=95, bottomMargin=155)
    styles = getSampleStyleSheet()
    pc = HexColor(PRIMARY)
    story = []

    _al = _para_align(lang)
    title_st = ParagraphStyle('T', parent=styles['Title'], fontSize=16,
                               textColor=pc, fontName=_fn, alignment=_al)
    sub_st   = ParagraphStyle('S', parent=styles['Normal'], fontSize=10,
                               textColor=HexColor('#5a7080'), fontName=_fn, alignment=_al)
    sec_st   = ParagraphStyle('Sec', parent=styles['Normal'], fontSize=11,
                               fontName=_fnb, textColor=pc,
                               spaceBefore=12, spaceAfter=6, alignment=_al)

    story.append(Paragraph(_c(f"{_t('Commission Summary')} — {period_label}"), title_st))
    story.append(Paragraph(_c(COMPANY), sub_st))
    story.append(HRFlowable(width="100%", thickness=2, color=pc))
    story.append(Spacer(1, 10))

    story.append(Paragraph(_c(_t("Company Summary")), sec_st))
    summary_data = [
        [_c(_t("Metric")), _c(_t("Value"))],
        [_c(_t("Total Sales (SAR)")),      f"SAR {totals['total_sales']:,.0f}"],
        [_c(_t("Total Target (SAR)")),     f"SAR {totals['total_target']:,.0f}"],
        [_c(_t("Achievement %")),          f"{totals['achievement']:.1f}%"],
        [_c(_t("Base Commission (SAR)")),  f"SAR {totals['total_base']:,.0f}"],
        [_c(_t("Final Commission (SAR)")), f"SAR {totals['total_final']:,.0f}"],
        [_c(_t("Total Salespersons")),     str(totals['total_count'])],
        [_c(_t("Achieved Target")),        str(totals['achieved_count'])],
    ]
    stbl = Table(_rtl_rows(summary_data, lang),
                 colWidths=_rtl_widths([250, 180], lang), hAlign=_rtl_halign(lang))
    stbl.setStyle(TableStyle([
        ('BACKGROUND',     (0, 0), (-1, 0),  pc),
        ('TEXTCOLOR',      (0, 0), (-1, 0),  HexColor('#ffffff')),
        ('FONTNAME',       (0, 0), (-1, -1), _fn),
        ('FONTNAME',       (0, 0), (-1, 0),  _fnb),
        ('FONTSIZE',       (0, 0), (-1, -1), 9),
        ('GRID',           (0, 0), (-1, -1), 0.3, HexColor('#dde5ea')),
        ('PADDING',        (0, 0), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [HexColor('#ffffff'), HexColor('#f7f9fb')]),
    ] + _rtl_style(lang)))
    story.append(stbl)
    story.append(Spacer(1, 14))

    story.append(Paragraph(_c(_t("All Salespersons")), sec_st))
    sp_data = [[_c(_t("Salesperson")), _c(_t("Branch")), _c(_t("Tier")),
                _c(_t("Total Sales (SAR)")), _c(_t("Achievement %")),
                _c(_t("KPI x")), _c(_t("Final Commission (SAR)"))]]
    for c in commissions:
        sp_data.append([
            _c(c['salesperson_name']), _c(c['branch_name'] or ''),
            _c(c['tier_name'] or ''),
            f"{c['total_actual']:,.0f}", f"{c['achievement']:.1f}%",
            f"x{c['kpi_multiplier']}", f"{c['final_commission']:,.0f}",
        ])
    sp_tbl = Table(_rtl_rows(sp_data, lang), hAlign=_rtl_halign(lang), repeatRows=1)
    sp_tbl.setStyle(TableStyle([
        ('BACKGROUND',     (0, 0), (-1, 0),  pc),
        ('TEXTCOLOR',      (0, 0), (-1, 0),  HexColor('#ffffff')),
        ('FONTNAME',       (0, 0), (-1, -1), _fn),
        ('FONTNAME',       (0, 0), (-1, 0),  _fnb),
        ('FONTSIZE',       (0, 0), (-1, -1), 7),
        ('GRID',           (0, 0), (-1, -1), 0.3, HexColor('#dde5ea')),
        ('PADDING',        (0, 0), (-1, -1), 4),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [HexColor('#ffffff'), HexColor('#f7f9fb')]),
    ] + _rtl_style(lang)))
    story.append(sp_tbl)

    doc.build(story, onFirstPage=_draw_page_bg, onLaterPages=_draw_page_bg)
    return buf.getvalue()


# ─── Branch PDF ───────────────────────────────────────────────────────────────
def make_pdf_branch(branch_name: str, persons: list, lang: str = 'en') -> bytes:
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.colors import HexColor
    from reportlab.lib.pagesizes import A4
    def _t(s): return tl(s, lang)
    _c = lambda s: _ar_cell(s, lang)
    _fn  = pdf_font(lang)
    _fnb = pdf_font_bold(lang)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=50, rightMargin=50,
                            topMargin=95, bottomMargin=155)
    styles = getSampleStyleSheet()
    pc = HexColor(PRIMARY)
    ac = HexColor(ACCENT)
    story = []

    _al = _para_align(lang)
    title_st = ParagraphStyle('T', parent=styles['Title'], fontSize=16,
                               textColor=pc, fontName=_fn, alignment=_al)
    sub_st   = ParagraphStyle('S', parent=styles['Normal'], fontSize=10,
                               textColor=HexColor('#5a7080'), fontName=_fn, alignment=_al)
    sec_st   = ParagraphStyle('Sec', parent=styles['Normal'], fontSize=11,
                               fontName=_fnb, textColor=pc,
                               spaceBefore=12, spaceAfter=6, alignment=_al)

    story.append(Paragraph(_c(f"{_t('Branch Report')} — {branch_name} | {period_label}"), title_st))
    story.append(Paragraph(_c(COMPANY), sub_st))
    story.append(HRFlowable(width="100%", thickness=2, color=pc))
    story.append(Spacer(1, 10))

    story.append(Paragraph(_c(_t("All Salespersons")), sec_st))
    tbl_data = [[_c(_t("Salesperson")), _c(_t("Tier")),
                 _c(_t("Total Sales (SAR)")), _c(_t("Total Target (SAR)")),
                 _c(_t("Achievement %")), _c(_t("KPI x")),
                 _c(_t("Final Commission (SAR)"))]]
    b_sales = b_target = b_final = 0.0
    for c in persons:
        ach = (c['total_actual'] / c['total_target'] * 100) if c['total_target'] else 0
        tbl_data.append([
            _c(c['salesperson_name']), _c(c['tier_name'] or ''),
            f"{c['total_actual']:,.0f}", f"{c['total_target']:,.0f}",
            f"{ach:.1f}%", f"x{c['kpi_multiplier']}",
            f"{c['final_commission']:,.0f}",
        ])
        b_sales  += c['total_actual']
        b_target += c['total_target']
        b_final  += c['final_commission']
    tbl_data.append([
        _c(_t("TOTAL")), "",
        f"{b_sales:,.0f}", f"{b_target:,.0f}",
        f"{(b_sales/b_target*100) if b_target else 0:.1f}%", "",
        f"{b_final:,.0f}",
    ])

    tbl = Table(_rtl_rows(tbl_data, lang), hAlign=_rtl_halign(lang), repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND',     (0, 0),  (-1, 0),  pc),
        ('TEXTCOLOR',      (0, 0),  (-1, 0),  HexColor('#ffffff')),
        ('FONTNAME',       (0, 0),  (-1, -1), _fn),
        ('FONTNAME',       (0, 0),  (-1, 0),  _fnb),
        ('BACKGROUND',     (0, -1), (-1, -1), ac),
        ('FONTNAME',       (0, -1), (-1, -1), _fnb),
        ('FONTSIZE',       (0, 0),  (-1, -1), 8),
        ('GRID',           (0, 0),  (-1, -1), 0.3, HexColor('#dde5ea')),
        ('PADDING',        (0, 0),  (-1, -1), 5),
        ('ROWBACKGROUNDS', (0, 1),  (-1, -2),
         [HexColor('#ffffff'), HexColor('#f7f9fb')]),
    ] + _rtl_style(lang)))
    story.append(tbl)

    doc.build(story, onFirstPage=_draw_page_bg, onLaterPages=_draw_page_bg)
    return buf.getvalue()


# ─── Salesperson PDF ──────────────────────────────────────────────────────────
def make_pdf_salesperson(c: dict, lang: str = 'en') -> bytes:
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable,
                                    Image as _RLImage)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.colors import HexColor
    from reportlab.lib.pagesizes import A4
    def _t(s): return tl(s, lang)
    _c = lambda s: _ar_cell(s, lang)
    _fn  = pdf_font(lang)
    _fnb = pdf_font_bold(lang)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=50, rightMargin=50,
                            topMargin=95, bottomMargin=155)
    styles = getSampleStyleSheet()
    pc = HexColor(PRIMARY)
    ac = HexColor(ACCENT)
    story = []

    _al = _para_align(lang)
    title_st = ParagraphStyle('T', parent=styles['Title'],
                               fontSize=16, textColor=pc, fontName=_fn, alignment=_al)
    sub_st   = ParagraphStyle('S', parent=styles['Normal'],
                               fontSize=10, textColor=HexColor('#5a7080'),
                               fontName=_fn, alignment=_al)
    sec_st   = ParagraphStyle('Sec', parent=styles['Normal'],
                               fontSize=11, fontName=_fnb,
                               textColor=pc, spaceBefore=12, spaceAfter=6, alignment=_al)

    # Header — logo and title (sides swap for RTL)
    from reportlab.platypus import Table as RLTable, TableStyle as RLTS
    from reportlab.lib.units import cm
    header_logo = None
    if _LOGO_BYTES:
        try:
            from reportlab.lib.utils import ImageReader
            header_logo = _RLImage(ImageReader(io.BytesIO(_LOGO_BYTES)),
                                   width=2.4 * cm, height=1.2 * cm)
        except Exception:
            header_logo = None
    title_block = [
        Paragraph(_c(COMPANY), title_st),
        Paragraph(
            _c(f"{_t('Commission Summary')} | {period_label} | "
               f"{c['salesperson_name']} | {c['branch_name']}"),
            sub_st,
        ),
    ]
    if header_logo:
        if _is_ar(lang):   # logo on the right for RTL
            hdr_cells, hdr_widths = [[title_block, header_logo]], [None, 2.8 * cm]
        else:
            hdr_cells, hdr_widths = [[header_logo, title_block]], [2.8 * cm, None]
        hdr_tbl = RLTable(hdr_cells, colWidths=hdr_widths)
        hdr_tbl.setStyle(RLTS([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (0, 0), 0),
        ]))
        story.append(hdr_tbl)
    else:
        story.extend(title_block)
    story.append(HRFlowable(width="100%", thickness=2, color=pc))
    story.append(Spacer(1, 10))

    story.append(Paragraph(_c(_t("Category Breakdown")), sec_st))
    tbl_data = [[_c(_t("Category")), _c(_t("Actual Sales")), _c(_t("Target")),
                 _c(_t("Achievement %")), _c(_t("Rate %")), _c(_t("Commission"))]]
    for cr in c['categories']:
        ach = (cr['actual_sales'] / cr['target'] * 100) if cr['target'] else 0
        tbl_data.append([
            _c(cr['category_name']),
            f"SAR {cr['actual_sales']:,.0f}",
            f"SAR {cr['target']:,.0f}",
            f"{ach:.1f}%",
            f"{cr['rate']:.2f}%",
            f"SAR {cr['commission']:,.0f}",
        ])
    tbl = Table(_rtl_rows(tbl_data, lang), hAlign=_rtl_halign(lang), repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND',   (0, 0), (-1, 0),  pc),
        ('TEXTCOLOR',    (0, 0), (-1, 0),  HexColor('#ffffff')),
        ('FONTNAME',     (0, 0), (-1, -1), _fn),
        ('FONTNAME',     (0, 0), (-1, 0),  _fnb),
        ('FONTSIZE',     (0, 0), (-1, -1), 8),
        ('GRID',         (0, 0), (-1, -1), 0.3, HexColor('#dde5ea')),
        ('PADDING',      (0, 0), (-1, -1), 5),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [HexColor('#ffffff'), HexColor('#f7f9fb')]),
    ] + _rtl_style(lang)))
    story.append(tbl)
    story.append(Spacer(1, 12))

    # ── KPI Breakdown — each KPI item with its score, weight and contribution ──
    kpi = c.get('kpi') or {}
    kpi_items = kpi.get('items') or []
    story.append(Paragraph(_c(_t("KPI Score Breakdown")), sec_st))
    if not kpi_items:
        story.append(Paragraph(_c(_t("No KPI items configured.")), sub_st))
    else:
        kpi_data = [[_c(_t("KPI Item")), _c(_t("Weight %")), _c(_t("Score")),
                     _c(_t("Achieved %")), _c(_t("Weighted Points")), _c(_t("Source"))]]
        for it in kpi_items:
            src = _t("Auto") if it.get('is_auto') else _t("Manual")
            kpi_data.append([
                _c(it['name']),
                f"{it['weight']:.0f}%",
                f"{it['raw_score']:.0f} / {it['max_score']:.0f}",
                f"{it['normalized']:.1f}%",
                f"{it['contribution']:.1f}",
                _c(src),
            ])
        # summary rows
        kpi_data.append([_c(_t("Weighted Score")), "", "", "",
                         f"{kpi.get('weighted_score', 0):.1f}", ""])
        if kpi.get('bonus'):
            kpi_data.append([_c(_t("Bonus")), "", "", "",
                             f"+{kpi['bonus']:.1f}", ""])
        if kpi.get('penalty'):
            kpi_data.append([_c(_t("Penalty")), "", "", "",
                             f"-{kpi['penalty']:.1f}", ""])
        kpi_data.append([_c(_t("Final KPI Score")), "", "", "",
                         f"{kpi.get('final_score', c['kpi_score']):.1f}", ""])
        # Applied multiplier rule → shows HOW the multiplier was derived
        _rule = kpi.get('applied_rule')
        if _rule:
            _hi = "∞" if (_rule.get('is_unlimited') or _rule.get('score_to') is None) else f"{_rule['score_to']:.0f}"
            _rng = f"{_rule['score_from']:.0f}–{_hi}"
        else:
            _rng = "—"
        kpi_data.append([_c(_t("KPI Multiplier")), "", "", "",
                         f"x {c['kpi_multiplier']}", _c(_rng)])
        # summary rows = Weighted Score, Final KPI Score, KPI Multiplier (+ bonus/penalty)
        n_summary = 3 + (1 if kpi.get('bonus') else 0) + (1 if kpi.get('penalty') else 0)
        ktbl = Table(_rtl_rows(kpi_data, lang), hAlign=_rtl_halign(lang), repeatRows=1)
        ktbl.setStyle(TableStyle([
            ('BACKGROUND',   (0, 0), (-1, 0),  pc),
            ('TEXTCOLOR',    (0, 0), (-1, 0),  HexColor('#ffffff')),
            ('FONTNAME',     (0, 0), (-1, -1), _fn),
            ('FONTNAME',     (0, 0), (-1, 0),  _fnb),
            ('FONTNAME',     (0, -1), (-1, -1), _fnb),
            ('BACKGROUND',   (0, -1), (-1, -1), ac),
            ('FONTSIZE',     (0, 0), (-1, -1), 8),
            ('GRID',         (0, 0), (-1, -1), 0.3, HexColor('#dde5ea')),
            ('PADDING',      (0, 0), (-1, -1), 5),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1 - n_summary),
             [HexColor('#ffffff'), HexColor('#f7f9fb')]),
        ] + _rtl_style(lang)))
        story.append(ktbl)
    story.append(Spacer(1, 12))

    story.append(Paragraph(_c(_t("Commission Summary")), sec_st))
    summary = [
        [_c(_t("Base Commission")), f"SAR {c['base_commission']:,.0f}"],
        [_c(_t("KPI Score")),       f"{c['kpi_score']:.2f}"],
        [_c(_t("KPI Multiplier")),  f"x {c['kpi_multiplier']}"],
        [_c(_t("FINAL COMMISSION")), f"SAR {c['final_commission']:,.0f}"],
    ]
    stbl = Table(_rtl_rows(summary, lang), colWidths=_rtl_widths([200, 200], lang),
                 hAlign=_rtl_halign(lang))
    stbl.setStyle(TableStyle([
        ('FONTNAME',    (0, 0), (-1, -1), _fn),
        ('FONTSIZE',    (0, 0), (-1, -1), 9),
        ('FONTNAME',    (0, -1), (-1, -1), _fnb),
        ('FONTSIZE',    (0, -1), (-1, -1), 11),
        ('BACKGROUND',  (0, -1), (-1, -1), ac),
        ('TEXTCOLOR',   (0, -1), (-1, -1), pc),
        ('GRID',        (0, 0), (-1, -1), 0.3, HexColor('#dde5ea')),
        ('PADDING',     (0, 0), (-1, -1), 6),
    ] + _rtl_style(lang)))
    story.append(stbl)

    doc.build(story, onFirstPage=_draw_page_bg, onLaterPages=_draw_page_bg)
    return buf.getvalue()


# ── Reusable download options widget ─────────────────────────────────────────
def _download_options(key_prefix: str) -> tuple:
    """Render language + format selectors. Returns (lang_code, is_pdf)."""
    st.markdown(
        f"<div style='font-size:12px;font-weight:600;color:#354f61;"
        f"text-transform:uppercase;letter-spacing:.06em;margin-bottom:6px'>"
        f"{t('Report Options')}</div>",
        unsafe_allow_html=True,
    )
    oc1, oc2 = st.columns(2)
    with oc1:
        lang_sel = st.radio(
            t("Report Language"),
            ["English", "العربية"],
            index=1 if is_rtl() else 0,
            horizontal=False,
            key=f"{key_prefix}_lang",
        )
    with oc2:
        fmt_sel = st.radio(
            t("Format"),
            ["Excel", "PDF"],
            horizontal=False,
            key=f"{key_prefix}_fmt",
        )
    _lang = 'ar' if 'العربية' in lang_sel else 'en'
    _pdf  = fmt_sel == "PDF"
    return _lang, _pdf


# ══════════════════════════════════════════════════════════════════════════════
# THREE-LEVEL REPORT TABS
# ══════════════════════════════════════════════════════════════════════════════
tab_company, tab_region, tab_branch, tab_person = st.tabs([
    t("Company Report"), t("Region Report"), t("Branch Report"), t("Salesperson Report"),
])

# ────────────────────────────────────────────────────────────────────────────
# COMPANY REPORT
# ────────────────────────────────────────────────────────────────────────────
with tab_company:
    st.markdown(f"### {t('Company-Wide Report')} — {period_label}")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric(t("Total Sales"),      f"SAR {totals['total_sales']:,.0f}")
    c2.metric(t("Achievement"),      f"{totals['achievement']:.1f}%")
    c3.metric(t("Base Commission"),  f"SAR {totals['total_base']:,.0f}")
    c4.metric(t("Final Commission"), f"SAR {totals['total_final']:,.0f}")

    st.markdown("---")
    _dl_lang_co, _dl_pdf_co = _download_options("co")
    if _dl_pdf_co:
        st.download_button(
            label=t('Download PDF'),
            data=make_pdf_company(lang=_dl_lang_co),
            file_name=f"{COMPANY}_{period_label.replace(' ','_')}_Report.pdf",
            mime="application/pdf",
            type="primary", use_container_width=True,
        )
    else:
        st.download_button(
            label=t('Download Excel'),
            data=make_excel_company(lang=_dl_lang_co),
            file_name=f"{COMPANY}_{period_label.replace(' ','_')}_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary", use_container_width=True,
        )
    st.caption(t("Includes: Company Summary, All Salespersons, By Branch, Category Breakdown"))


# ────────────────────────────────────────────────────────────────────────────
# REGION REPORT
# ────────────────────────────────────────────────────────────────────────────
with tab_region:
    st.markdown(f"### {t('Region Reports')} — {period_label}")

    region_groups: dict = {}
    for c in commissions:
        reg = _branch_to_region.get(c['branch_name'] or '', '') or t("No Region")
        region_groups.setdefault(reg, []).append(c)

    if not region_groups:
        st.info(t("No commission data for this period."))
    else:
        # Summary table
        re_rows = []
        for reg_name, persons in region_groups.items():
            r_sales  = sum(c['total_actual'] for c in persons)
            r_target = sum(c['total_target'] for c in persons)
            r_final  = sum(c['final_commission'] for c in persons)
            ach = (r_sales / r_target * 100) if r_target else 0
            re_rows.append({
                t("Region"):           reg_name,
                t("Salespersons"):     len(persons),
                t("Total Sales"):      f"SAR {r_sales:,.0f}",
                t("Target"):           f"SAR {r_target:,.0f}",
                t("Achievement"):      f"{ach:.1f}%",
                t("Final Commission"): f"SAR {r_final:,.0f}",
            })
        render_df(pd.DataFrame(re_rows))
        st.markdown("---")

        # Per-region download
        sel_reg = st.selectbox(t("Select region to download"), list(region_groups.keys()),
                               key="rep_re_sel")
        if sel_reg:
            reg_persons = region_groups[sel_reg]
            st.markdown(f"**{t('Region Report')}: {sel_reg}**")
            _dl_lang_re, _dl_pdf_re = _download_options("re")
            if _dl_pdf_re:
                st.download_button(
                    label=t('Download PDF'),
                    data=make_pdf_branch(sel_reg, reg_persons, lang=_dl_lang_re),
                    file_name=f"Region_{sel_reg}_{period_label.replace(' ','_')}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
            else:
                st.download_button(
                    label=f"{t('Download Excel')} — {sel_reg}",
                    data=make_excel_branch(sel_reg, reg_persons, lang=_dl_lang_re),
                    file_name=f"Region_{sel_reg}_{period_label.replace(' ','_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )


# ────────────────────────────────────────────────────────────────────────────
# BRANCH REPORT
# ────────────────────────────────────────────────────────────────────────────
with tab_branch:
    st.markdown(f"### {t('Branch Reports')} — {period_label}")

    branch_groups: dict = {}
    for c in commissions:
        b = c['branch_name'] or 'Unknown'
        branch_groups.setdefault(b, []).append(c)

    if not branch_groups:
        st.info(t("No commission data for this period."))
    else:
        # Summary table
        br_rows = []
        for br_name, persons in branch_groups.items():
            b_sales  = sum(c['total_actual'] for c in persons)
            b_target = sum(c['total_target'] for c in persons)
            b_final  = sum(c['final_commission'] for c in persons)
            ach = (b_sales / b_target * 100) if b_target else 0
            br_rows.append({
                t("Branch"):           br_name,
                t("Region"):           _branch_to_region.get(br_name, '') or '—',
                t("Salespersons"):     len(persons),
                t("Total Sales"):      f"SAR {b_sales:,.0f}",
                t("Target"):           f"SAR {b_target:,.0f}",
                t("Achievement"):      f"{ach:.1f}%",
                t("Final Commission"): f"SAR {b_final:,.0f}",
            })
        render_df(pd.DataFrame(br_rows))
        st.markdown("---")

        # Per-branch download
        sel_br = st.selectbox(t("Select branch to download"), list(branch_groups.keys()),
                              key="rep_br_sel")
        if sel_br:
            br_persons = branch_groups[sel_br]
            st.markdown(f"**{t('Branch Report')}: {sel_br}**")
            _dl_lang_br, _dl_pdf_br = _download_options("br")
            if _dl_pdf_br:
                st.download_button(
                    label=t('Download PDF'),
                    data=make_pdf_branch(sel_br, br_persons, lang=_dl_lang_br),
                    file_name=f"{sel_br}_{period_label.replace(' ','_')}_Branch.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
            else:
                st.download_button(
                    label=t('Download {branch} — Excel').format(branch=sel_br),
                    data=make_excel_branch(sel_br, br_persons, lang=_dl_lang_br),
                    file_name=f"{sel_br}_{period_label.replace(' ','_')}_Branch.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

        st.markdown("---")
        st.markdown(f"**{t('Download All Branches (one file)')}**")
        _dl_lang_all, _dl_pdf_all = _download_options("br_all")
        if _dl_pdf_all:
            st.download_button(
                label=t('Download PDF'),
                data=make_pdf_company(lang=_dl_lang_all),
                file_name=f"AllBranches_{period_label.replace(' ','_')}_Report.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
        else:
            st.download_button(
                label=t('Download All Branches (one file)'),
                data=make_excel_company(lang=_dl_lang_all),
                file_name=f"AllBranches_{period_label.replace(' ','_')}_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )


# ────────────────────────────────────────────────────────────────────────────
# SALESPERSON REPORT
# ────────────────────────────────────────────────────────────────────────────
with tab_person:
    st.markdown(f"### {t('Salesperson Reports')} — {period_label}")

    if not commissions:
        st.info(t("No commission data for this period."))
    else:
        sp_names    = [c['salesperson_name'] for c in commissions]
        selected_sp = st.selectbox(t("Select Salesperson"), sp_names, key="rep_sp_sel")
        sel_c       = next((c for c in commissions if c['salesperson_name'] == selected_sp), None)

        if sel_c:
            st.markdown(
                f"**{sel_c['salesperson_name']}** | "
                f"{sel_c['branch_name']} | {sel_c['tier_name']}"
            )
            m1, m2, m3, m4 = st.columns(4)
            m1.metric(t("Total Sales"),  f"SAR {sel_c['total_actual']:,.0f}")
            m2.metric(t("Achievement"),  f"{sel_c['achievement']:.1f}%")
            m3.metric(t("Base Comm."),   f"SAR {sel_c['base_commission']:,.0f}")
            m4.metric(t("Final Comm."),  f"SAR {sel_c['final_commission']:,.0f}")

            # ── KPI Score Breakdown ───────────────────────────────────────────
            _kpi    = sel_c.get('kpi') or {}
            _kitems = _kpi.get('items') or []
            st.markdown(f"#### {t('KPI Score Breakdown')}")
            if not _kitems:
                st.caption(t("No KPI items configured."))
            else:
                _kpi_rows = [{
                    t("KPI Item"):         it['name'],
                    t("Weight %"):         f"{it['weight']:.0f}%",
                    t("Score"):            f"{it['raw_score']:.0f} / {it['max_score']:.0f}",
                    t("Achieved %"):       f"{it['normalized']:.1f}%",
                    t("Weighted Points"):  f"{it['contribution']:.1f}",
                    t("Source"):           t("Auto") if it.get('is_auto') else t("Manual"),
                } for it in _kitems]
                render_df(pd.DataFrame(_kpi_rows))
                k1, k2, k3, k4 = st.columns(4)
                k1.metric(t("Weighted Score"),  f"{_kpi.get('weighted_score', 0):.1f}")
                if _kpi.get('bonus') or _kpi.get('penalty'):
                    k2.metric(t("Bonus / Penalty"),
                              f"+{_kpi.get('bonus', 0):.0f} / -{_kpi.get('penalty', 0):.0f}")
                k3.metric(t("Final KPI Score"),
                          f"{_kpi.get('final_score', sel_c['kpi_score']):.1f}")
                _rule = _kpi.get('applied_rule')
                if _rule:
                    _hi = "∞" if (_rule.get('is_unlimited') or _rule.get('score_to') is None) else f"{_rule['score_to']:.0f}"
                    _mhelp = f"{t('Rule')}: {_rule['score_from']:.0f}–{_hi}"
                else:
                    _mhelp = None
                k4.metric(t("KPI Multiplier"), f"x {sel_c['kpi_multiplier']}", help=_mhelp)

            st.markdown("---")
            _dl_lang_sp, _dl_pdf_sp = _download_options("sp")
            if _dl_pdf_sp:
                st.download_button(
                    label=t('Download PDF'),
                    data=make_pdf_salesperson(sel_c, lang=_dl_lang_sp),
                    file_name=f"{selected_sp.replace(' ','_')}_{period_label.replace(' ','_')}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
            else:
                st.download_button(
                    label=t('Download Excel'),
                    data=make_excel_salesperson(sel_c, lang=_dl_lang_sp),
                    file_name=f"{selected_sp.replace(' ','_')}_{period_label.replace(' ','_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

        st.divider()
        st.markdown(f"#### {t('All Salespersons Quick View')}")
        quick_rows = [{
            t("Name"):        c['salesperson_name'],
            t("Branch"):      c['branch_name'],
            t("Sales"):       f"SAR {c['total_actual']:,.0f}",
            t("Achievement"): f"{c['achievement']:.1f}%",
            t("Final Comm."): f"SAR {c['final_commission']:,.0f}",
        } for c in sorted(commissions, key=lambda x: x['final_commission'], reverse=True)]
        render_df(pd.DataFrame(quick_rows))


