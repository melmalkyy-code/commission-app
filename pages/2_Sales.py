import sys, os; sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
import streamlit as st
from src.startup import init_db
from src.auth import require_login, require_editor
init_db()
require_login()
require_editor()   # viewers are read-only — no sales input
import pandas as pd
from src.models import (get_setting, get_salespersons, get_categories,
                         get_or_create_period, save_sales_bulk, get_sales,
                         get_tier_target, log_action)
from src.calculations import calc_all_commissions, get_totals
from src.ui import inject_css, page_header, sidebar_logo, render_df
from src.i18n import t, q_label
from src.branding import page_icon

PRIMARY = get_setting('primary_color', '#354f61')
ACCENT  = get_setting('accent_color', '#f6ba3b')
COMPANY = get_setting('company_name', 'Surveying Experts')

st.set_page_config(page_title="Sales Input — Surveying Experts", page_icon=page_icon(), layout="wide")
inject_css(PRIMARY)
sidebar_logo(COMPANY, PRIMARY)
page_header(t("Sales Input"), t("Enter actual sales amounts (SAR) for each salesperson"), PRIMARY)

col1, col2 = st.columns([1, 1])
year    = col1.selectbox(t("Year"),    [2024, 2025, 2026, 2027], index=2)
quarter = col2.selectbox(t("Quarter"), [1, 2, 3, 4],             index=1,
                          format_func=q_label)
period  = get_or_create_period(year, quarter)

if period.get('is_locked'):
    st.warning(t("This quarter is locked. Contact your manager to unlock it in Settings."))
    st.stop()

st.divider()

salespeople = get_salespersons(active_only=True)
categories  = [c for c in get_categories(active_only=True) if c['include_in_commission']]

if not salespeople:
    st.info(t("No active salespersons found. Please add salespersons in Settings first."))
    st.stop()
if not categories:
    st.info(t("No active categories found. Please configure categories in Settings first."))
    st.stop()

def _read_row_val(row, col_name: str) -> float:
    """Read a cell value by column name, returning 0.0 on null/blank."""
    raw = row.get(col_name)
    try:
        return 0.0 if (raw is None or pd.isna(raw)) else float(raw)
    except Exception:
        return 0.0


period_id    = period['id']
display_cols = [t('Salesperson'), t('Branch'), t('Tier')] + [c['name'] for c in categories]

# A signature that changes if the period, salespersons or categories change,
# so the editor rebuilds only when the underlying set actually changes.
_sig      = f"{period_id}-{'.'.join(str(s['id']) for s in salespeople)}-{'.'.join(str(c['id']) for c in categories)}"
_base_key = f"_sales_base_{_sig}"
_snap_key = f"_sales_snap_{_sig}"

# Build the editor's DataFrame + saved-value snapshot ONCE per signature.
# It then lives in session_state and is NOT rebuilt from the DB on every
# rerun — that is what previously wiped the value you were typing.
if _base_key not in st.session_state:
    db_sales  = get_sales(period_id)
    sales_map = {(r['salesperson_id'], r['category_id']): float(r['actual_sales'] or 0)
                 for r in db_sales}
    rows = []
    for sp in salespeople:
        row = {t('Salesperson'): sp['name'],
               t('Branch'): sp.get('branch_name') or '',
               t('Tier'):   sp.get('tier_name') or ''}
        for cat in categories:
            row[cat['name']] = float(sales_map.get((sp['id'], cat['id']), 0.0))
        rows.append(row)
    st.session_state[_base_key] = pd.DataFrame(rows)[display_cols]
    st.session_state[_snap_key] = {
        (sp['id'], cat['id']): float(sales_map.get((sp['id'], cat['id']), 0.0))
        for sp in salespeople for cat in categories
    }

st.markdown(f"**{t('Enter actual sales amounts (SAR) for each salesperson:')}**")
st.caption(t("Changes are saved automatically as you edit each cell."))

edited = st.data_editor(
    st.session_state[_base_key],
    use_container_width=True,
    disabled=[t('Salesperson'), t('Branch'), t('Tier')],
    num_rows="fixed",
    column_config={
        cat['name']: st.column_config.NumberColumn(
            cat['name'], min_value=0, format="SAR %d", step=1000,
        ) for cat in categories
    },
    key=f"sales_editor_{_sig}",
    hide_index=True,
)

# ── Fast autosave: only the cells that actually changed, in ONE round-trip ────
snapshot = st.session_state[_snap_key]
to_save  = []
for idx, sp in enumerate(salespeople):
    if idx >= len(edited):
        break
    row = edited.iloc[idx]
    for cat in categories:
        newv = _read_row_val(row, cat['name'])
        key  = (sp['id'], cat['id'])
        if newv != snapshot.get(key, 0.0):
            to_save.append((sp['id'], cat['id'], newv))

if to_save:
    save_sales_bulk(period_id, to_save)
    for sp_id, cat_id, v in to_save:
        snapshot[(sp_id, cat_id)] = v
    # Invalidate the (cached) commission calc so dashboards/reports refresh.
    # This is a cheap cache-drop — it does NOT recompute here.
    calc_all_commissions.clear()
    st.toast(f"{t('Saved')} · {len(to_save)}")

# ── Instant running total (from the snapshot, no database call) ───────────────
st.divider()
_entered = sum(snapshot.values())
mc1, mc2 = st.columns([1, 3])
mc1.metric(t("Total Sales Entered"), f"SAR {_entered:,.0f}")

# ── Commission preview — computed on demand so typing stays instant ───────────
if st.button(t("Calculate Commission Preview"), type="primary"):
    calc_all_commissions.clear()
    with st.spinner(t("Calculating...")):
        commissions = calc_all_commissions(period_id)
        totals      = get_totals(commissions)
    st.session_state['_sales_preview'] = {
        'totals': totals,
        'rows': [{
            t("Salesperson"):  c['salesperson_name'],
            t("Branch"):       c['branch_name'] or '',
            t("Total Sales"):  f"SAR {c['total_actual']:,.0f}",
            t("Target"):       f"SAR {c['total_target']:,.0f}",
            t("Achievement"):  f"{c['achievement']:.1f}%",
            t("Base Comm."):   f"SAR {c['base_commission']:,.0f}",
            t("KPI Mult."):    f"x {c['kpi_multiplier']}",
            t("Final Comm."):  f"SAR {c['final_commission']:,.0f}",
        } for c in commissions],
    }

if st.session_state.get('_sales_preview'):
    _pv = st.session_state['_sales_preview']
    _tt = _pv['totals']
    st.markdown(f"#### {t('Commission Preview')}")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric(t("Total Sales"),      f"SAR {_tt['total_sales']:,.0f}")
    c2.metric(t("Total Target"),     f"SAR {_tt['total_target']:,.0f}")
    c3.metric(t("Achievement"),      f"{_tt['achievement']:.1f}%")
    c4.metric(t("Final Commission"), f"SAR {_tt['total_final']:,.0f}")
    render_df(pd.DataFrame(_pv['rows']))

# ── Import / Export ────────────────────────────────────────────────────────────
st.divider()
with st.expander(t("Import from Excel / Download Template")):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Import"
    headers = ["Year", "Quarter", "Salesperson Name"] + [c['name'] for c in categories]
    fill = PatternFill(start_color="354f61", end_color="354f61", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
    for sp in salespeople:
        ws.append([year, quarter, sp['name']] + [0] * len(categories))

    buf = io.BytesIO()
    wb.save(buf)
    st.download_button(
        t("Download Import Template"), buf.getvalue(),
        file_name=f"Sales_Template_Q{quarter}_{year}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    uploaded = st.file_uploader(t("Upload filled template:"), type=['xlsx'])
    if uploaded:
        import openpyxl
        wb2 = openpyxl.load_workbook(uploaded)
        ws2 = wb2.active
        hdrs = [str(c.value).strip() if c.value else '' for c in ws2[1]]
        sp_map_by_name = {s['name']: s for s in salespeople}
        cat_map = {c['name']: c for c in categories}
        errors, to_import = [], []
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            d = dict(zip(hdrs, row))
            sp_name = str(d.get('Salesperson Name', '') or '').strip()
            sp = sp_map_by_name.get(sp_name)
            if not sp:
                errors.append(f"Salesperson not found: {sp_name}")
                continue
            for cat_name, cat in cat_map.items():
                val = d.get(cat_name)
                if val is not None:
                    try:
                        to_import.append((sp['id'], cat['id'], float(val)))
                    except Exception as e:
                        errors.append(str(e))
        if to_import:
            save_sales_bulk(period['id'], to_import)
        if errors:
            st.warning("\n".join(errors[:5]))
        # Drop the cached editor state so it rebuilds with the imported values
        for _k in [k for k in st.session_state
                   if k.startswith('_sales_base_') or k.startswith('_sales_snap_')]:
            del st.session_state[_k]
        calc_all_commissions.clear()
        st.success(t("Imported {count} records.").format(count=len(to_import)))
        st.rerun()


